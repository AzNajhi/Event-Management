import openpyxl
import os

def save(session_list, days, sessions):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for session in range(sessions):
        sheet.cell(row=1, column=session+2, value=f"Session {session+1}")

    for x in range(days):
        sheet.cell(row=x+2, column=1, value=f"Day {x+1}")
        for y in range(sessions):
            assigned_list = session_list[x][y].assigned_staff
            concat = ", ".join([f"{z}"for z in assigned_list])
            sheet.cell(row=x+2, column=y+2, value=concat)

    folder_name = 'autoplan_savefolder'

    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    counter = 0
    filename = 'autoplan_savefile'
    while os.path.exists(f"{folder_name}/{filename}.xlsx"):
        print("filename already existed!")
        counter += 1
        new_filename = f'{filename}_{counter}.xlsx'
        while os.path.exists(f"{folder_name}/{new_filename}"):
            print("new filename already existed!")
            counter += 1
            new_filename = f'{filename}_{counter}.xlsx'
        if not os.path.exists(f"{folder_name}/{new_filename}"):
            file_path = os.path.join(folder_name, new_filename)
            workbook.save(file_path)
            print("new file saved")
            return


    if not os.path.exists(f"{folder_name}/{filename}"):
        file_path = os.path.join(folder_name, f"{filename}.xlsx")
        workbook.save(file_path)
        print("file saved")
        return