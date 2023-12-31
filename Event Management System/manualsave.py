from tkinter import messagebox
import openpyxl
import os

def save(session_list, days, sessions):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for session in range(sessions):
        sheet.cell(row=1, column=session+2, value=f"Session {session+1}")

    for x in range(days):
        sheet.cell(row=x+2, column=1, value=f"Day {x+1}")
        for y in range(sessions):
            entry_list = session_list[x][y].form_dict["entry_list"]
            assigned_list = []
            for z in range(len(entry_list)):
                assigned_list.append(entry_list[z].get())
            concat = ", ".join([f"{z}"for z in assigned_list])
            sheet.cell(row=x+2, column=y+2, value=concat)

    folder_name = 'manuplan_savefolder'

    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    counter = 0
    filename = 'manuplan_savefile'
    while os.path.exists(f"{folder_name}/{filename}.xlsx"):
        counter += 1
        new_filename = f'{filename}_{counter}.xlsx'
        while os.path.exists(f"{folder_name}/{new_filename}"):
            counter += 1
            new_filename = f'{filename}_{counter}.xlsx'
        if not os.path.exists(f"{folder_name}/{new_filename}"):
            file_path = os.path.join(folder_name, new_filename)
            workbook.save(file_path)
            messagebox.showinfo("Message", "Schedule has been saved.")
            return

    if not os.path.exists(f"{folder_name}/{filename}"):
        file_path = os.path.join(folder_name, f"{filename}.xlsx")
        workbook.save(file_path)
        messagebox.showinfo("Message", "Schedule has been saved.")
        return